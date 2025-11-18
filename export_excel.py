import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

async def export_insights_to_excel(insights: list, user_id: int):
    """
    Экспорт инсайтов в Excel файл
    
    Args:
        insights: список инсайтов из БД
        user_id: ID пользователя (для имени файла)
    
    Returns:
        путь к созданному файлу
    """
    try:
        # Создаем новую Excel книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Инсайды"
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Заголовки
        headers = ["ID", "Дата создания", "Тема", "Описание", 
                   "Макрорегион", "Отрасль", "Файл прикреплен"]
        
        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Заполняем данными
        for row_num, insight in enumerate(insights, 2):
            # Форматируем дату
            created_at = insight.get('created_at', '')
            if isinstance(created_at, str):
                # Если дата в ISO формате, берем только дату
                created_at = created_at.split('T')[0] if 'T' in created_at else created_at
            
            row_data = [
                insight.get('id', ''),
                created_at,
                insight.get('theme', ''),
                insight.get('description', ''),
                insight.get('macro_region', ''),
                insight.get('industry', ''),
                "Да" if insight.get('file_id') else "Нет"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Выравнивание текста
                if col_num in [2, 6, 7]:  # центрируем дату, макрорегион, отрасль
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Настройка ширины столбцов
        column_widths = {
            'A': 8,      # ID
            'B': 15,     # Дата
            'C': 25,     # Тема
            'D': 40,     # Описание
            'E': 15,     # Макрорегион
            'F': 20,     # Отрасль
            'G': 15      # Файл
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Установка высоты строк
        ws.row_dimensions[1].height = 25  # высота заголовка
        
        # Автоматический перенос текста для всех ячеек
        for row in ws.iter_rows(min_row=2, max_row=len(insights) + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Сохранение файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"/tmp/insights_export_{user_id}_{timestamp}.xlsx"
        
        # Убедимся что директория существует
        os.makedirs(os.path.dirname(filename) or '.', exist_ok=True)
        
        wb.save(filename)
        logger.info(f"Excel file created: {filename}")
        
        return filename
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        raise

async def export_insights_to_excel_advanced(insights: list, user_id: int):
    """
    Расширенный экспорт с использованием pandas (если нужна большая обработка)
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Преобразуем в DataFrame
        df = pd.DataFrame(insights)
        
        # Переименовываем колонки
        df = df.rename(columns={
            'id': 'ID',
            'created_at': 'Дата создания',
            'theme': 'Тема',
            'description': 'Описание',
            'macro_region': 'Макрорегион',
            'industry': 'Отрасль',
            'file_id': 'Файл',
            'filename': 'Имя файла',
            'user_id': 'ID пользователя'
        })
        
        # Выбираем только нужные колонки
        columns_to_keep = ['ID', 'Дата создания', 'Тема', 'Описание', 
                           'Макрорегион', 'Отрасль', 'Файл']
        df = df[[col for col in columns_to_keep if col in df.columns]]
        
        # Форматируем дату
        if 'Дата создания' in df.columns:
            df['Дата создания'] = pd.to_datetime(df['Дата создания']).dt.strftime('%Y-%m-%d')
        
        # Заменяем NaN на пустые строки
        df = df.fillna('')
        
        # Создаем индикатор наличия файла
        if 'Файл' in df.columns:
            df['Файл'] = df['Файл'].apply(lambda x: 'Да' if x else 'Нет')
        
        # Сохраняем в Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"/tmp/insights_advanced_{user_id}_{timestamp}.xlsx"
        
        os.makedirs(os.path.dirname(filename) or '.', exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Инсайды', index=False)
            
            # Форматирование через openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Инсайды']
            
            # Стиль заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоматическая подгонка ширины
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Advanced Excel file created: {filename}")
        return filename
        
    except ImportError:
        logger.warning("pandas not installed, using basic export instead")
        return await export_insights_to_excel(insights, user_id)
    except Exception as e:
        logger.error(f"Error in advanced export: {e}")
        raise
